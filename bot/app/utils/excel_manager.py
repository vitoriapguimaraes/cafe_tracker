import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import logging

logger = logging.getLogger(__name__)

class ExcelManager:
    """
    Gerencia a leitura e escrita na planilha Excel de controle de gastos.
    """
    def __init__(self, file_path):
        self.file_path = file_path
        self._initialize_excel()

    def _initialize_excel(self):
        """
        Cria a planilha se não existir e garante que as abas "Gastos" e "Categorias"
        e seus cabeçalhos estejam presentes.
        """
        try:
            # Tenta carregar o workbook
            wb = load_workbook(self.file_path)
        except FileNotFoundError:
            # Se não existir, cria um novo
            logger.info(f"Criando nova planilha em: {self.file_path}")
            wb = pd.ExcelWriter(self.file_path, engine='openpyxl').book
            
        # Garante a aba "Gastos"
        if "Gastos" not in wb.sheetnames:
            ws_gastos = wb.create_sheet("Gastos")
            gastos_cols = ["data", "categoria", "valor", "conta", "descricao", "pessoa", "mes", "ano"]
            ws_gastos.append(gastos_cols)
            logger.info("Aba 'Gastos' criada com cabeçalhos.")
        
        # Garante a aba "Categorias"
        if "Categorias" not in wb.sheetnames:
            ws_categorias = wb.create_sheet("Categorias")
            categorias_cols = ["categoria", "teto_mensal"]
            ws_categorias.append(categorias_cols)
            # Adiciona um exemplo de teto
            ws_categorias.append(["Mercado", 500.00])
            ws_categorias.append(["Lazer", 300.00])
            logger.info("Aba 'Categorias' criada com cabeçalhos e exemplos.")

        # Remove a aba padrão criada pelo openpyxl se existir
        if 'Sheet' in wb.sheetnames:
            std_sheet = wb['Sheet']
            wb.remove(std_sheet)

        # Salva o workbook
        wb.save(self.file_path)
        logger.info("Planilha Excel inicializada com sucesso.")

    def registrar_gasto(self, categoria: str, valor: float, conta: str, data_str: str, descricao: str, pessoa: str):
        """
        Registra um novo gasto na aba "Gastos".
        """
        try:
            # Converte a data para o formato datetime e extrai mês e ano
            data_obj = datetime.strptime(data_str, '%d/%m')
            # Assume o ano atual se não for fornecido (o bot é para 2025, mas a data é dia/mês)
            ano = datetime.now().year
            data_completa = data_obj.replace(year=ano).strftime('%d/%m/%Y')
            mes = data_obj.month
            
            novo_gasto = [
                data_completa,
                categoria.capitalize(),
                valor,
                conta.capitalize(),
                descricao,
                pessoa,
                mes,
                ano
            ]

            wb = load_workbook(self.file_path)
            ws_gastos = wb["Gastos"]
            
            # Adiciona a nova linha
            ws_gastos.append(novo_gasto)
            
            wb.save(self.file_path)
            logger.info(f"Gasto registrado: {novo_gasto}")
            return True
        except ValueError as e:
            logger.error(f"Erro ao registrar gasto (formato de data inválido ou outro erro): {e}")
            return False
        except Exception as e:
            logger.error(f"Erro inesperado ao registrar gasto: {e}")
            return False

    def calcular_limite_restante(self, categoria: str, mes: int, ano: int) -> float:
        """
        Calcula o limite restante para uma categoria em um determinado mês/ano.
        """
        try:
            # 1. Ler tetos
            wb = load_workbook(self.file_path)
            ws_categorias = wb["Categorias"]
            
            # Converte a aba Categorias para DataFrame para fácil consulta
            data = ws_categorias.values
            cols = next(data)
            df_categorias = pd.DataFrame(data, columns=cols)
            
            # Encontra o teto mensal
            teto_mensal = df_categorias[df_categorias['categoria'].str.lower() == categoria.lower()]['teto_mensal'].iloc[0]
            
            # 2. Ler gastos
            ws_gastos = wb["Gastos"]
            data = ws_gastos.values
            cols = next(data)
            df_gastos = pd.DataFrame(data, columns=cols)
            
            # Garante que as colunas de mês e ano são numéricas
            df_gastos['mes'] = pd.to_numeric(df_gastos['mes'], errors='coerce')
            df_gastos['ano'] = pd.to_numeric(df_gastos['ano'], errors='coerce')
            df_gastos['valor'] = pd.to_numeric(df_gastos['valor'], errors='coerce')
            
            # Filtra gastos pela categoria, mês e ano
            gastos_filtrados = df_gastos[
                (df_gastos['categoria'].str.lower() == categoria.lower()) &
                (df_gastos['mes'] == mes) &
                (df_gastos['ano'] == ano)
            ]
            
            # Soma os gastos
            total_gasto = gastos_filtrados['valor'].sum()
            
            # 3. Calcular limite restante
            limite_restante = teto_mensal - total_gasto
            
            return limite_restante
            
        except IndexError:
            logger.warning(f"Categoria '{categoria}' não encontrada na aba 'Categorias'. Retornando 0.0.")
            return 0.0
        except Exception as e:
            logger.error(f"Erro ao calcular limite restante: {e}")
            return 0.0

# Inicializa a planilha ao importar o módulo
# excel_manager = ExcelManager(os.getenv("EXCEL_FILE_PATH"))
